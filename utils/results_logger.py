"""
实验结果记录器 - 自动记录到Excel
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
from datetime import datetime
import json


class ExperimentLogger:
    """实验结果记录器"""
    
    def __init__(self, excel_path='results/experiment_results.xlsx'):
        """
        Args:
            excel_path: Excel文件保存路径
        """
        self.excel_path = Path(excel_path)
        self.excel_path.parent.mkdir(parents=True, exist_ok=True)
        
        # 初始化结果存储
        self.results = {
            'Summary': [],
            'Group_A_BaselineComparison': [],
            'Group_B_PartitionStrategy': [],
            'Group_C_ClientNumber': [],
            'Group_D_LearningRate': [],
            'Group_E_DifferentialPrivacy': []
        }
        
    def log_result(self, experiment_group, result_dict):
        """
        记录单个实验结果
        
        Args:
            experiment_group: 实验组名称 ('A', 'B', 'C', 'D', 'E')
            result_dict: 实验结果字典
        """
        # 添加到对应实验组
        group_key = f'Group_{experiment_group}_{self._get_group_name(experiment_group)}'
        if group_key in self.results:
            self.results[group_key].append(result_dict)
        
        # 同时添加到总览表
        self.results['Summary'].append(result_dict)
        
    def _get_group_name(self, group):
        """获取实验组完整名称"""
        names = {
            'A': 'BaselineComparison',
            'B': 'PartitionStrategy',
            'C': 'ClientNumber',
            'D': 'LearningRate',
            'E': 'DifferentialPrivacy'
        }
        return names.get(group, 'Unknown')
    
    def create_result_dict(self, dataset, partition_type, alpha, num_clients, 
                          learning_rate, epsilon, method, metrics, 
                          training_info, notes=''):
        """
        创建标准化的结果字典
        
        Args:
            dataset: 数据集名称
            partition_type: 划分方式 ('lda', 'label_skew', 'quantity_skew')
            alpha: LDA的α值 (如果是LDA划分)
            num_clients: 客户端数量
            learning_rate: 学习率
            epsilon: 差分隐私ε (如果适用)
            method: 方法名称
            metrics: 性能指标字典 {accuracy, precision, recall, f1, auc}
            training_info: 训练信息 {total_rounds, convergence_round, training_time, ...}
            notes: 备注
        
        Returns:
            结果字典
        """
        # 生成实验ID
        exp_id = self._generate_experiment_id(
            dataset, partition_type, alpha, num_clients, 
            learning_rate, epsilon, method
        )
        
        result = {
            # 实验标识
            'Experiment_ID': exp_id,
            'Timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            
            # 实验配置
            'Dataset': dataset,
            'Partition_Type': partition_type,
            'Alpha': alpha if partition_type == 'lda' else '-',
            'Num_Clients': num_clients,
            'Learning_Rate': float(learning_rate) if learning_rate is not None else 0.001,
            'Epsilon': epsilon if method == 'feddeproto' else '-',
            'Method': method,
            
            # 性能指标
            'Accuracy': metrics.get('accuracy', 0),
            'Precision': metrics.get('precision', 0),
            'Recall': metrics.get('recall', 0),
            'F1_Score': metrics.get('f1_score', metrics.get('f1', 0)),  # 兼容两种key
            'AUC': metrics.get('auc', 0),
            
            # 训练信息
            'Total_Rounds': training_info.get('total_rounds', 0),
            'Convergence_Round': training_info.get('convergence_round', 0),
            'Training_Time_Sec': training_info.get('training_time', 0),
            'Avg_Round_Time_Sec': training_info.get('avg_round_time', 0),
            'Final_Loss': training_info.get('final_loss', 0),
            
            # FedDeProto特有信息
            'Stage1_Rounds': training_info.get('stage1_rounds', '-'),
            'Stage2_Rounds': training_info.get('stage2_rounds', '-'),
            'Stopped_Clients': training_info.get('stopped_clients', '-'),
            
            # 其他
            'GPU_Used': training_info.get('gpu_used', 'Yes'),
            'Notes': notes
        }
        
        return result
    
    def _generate_experiment_id(self, dataset, partition_type, alpha, 
                                num_clients, lr, epsilon, method):
        """生成唯一的实验ID"""
        # 数据集缩写
        ds_abbr = {
            'australian': 'AUS',
            'german': 'GER',
            'xinwang': 'XIN',
            'uci': 'UCI'
        }
        
        # 划分方式缩写
        pt_abbr = {
            'lda': 'L',
            'label_skew': 'LS',
            'quantity_skew': 'QS'
        }
        
        # 方法缩写
        method_abbr = method[:6].upper()
        
        # 构建ID
        if partition_type == 'lda':
            partition_str = f"{pt_abbr[partition_type]}{str(alpha).replace('.', '')}"
        else:
            partition_str = pt_abbr.get(partition_type, 'UNK')
        
        exp_id = (f"{ds_abbr.get(dataset, 'UNK')}_"
                 f"{partition_str}_"
                 f"C{num_clients}_"
                 f"LR{str(lr).replace('.', '')}_"
                 f"E{str(epsilon).replace('.', '') if epsilon != '-' else 'NA'}_"
                 f"{method_abbr}")
        
        return exp_id
    
    def save_to_excel(self):
        """保存所有结果到Excel文件"""
        print(f"\n保存实验结果到 {self.excel_path}...")
        
        with pd.ExcelWriter(self.excel_path, engine='openpyxl') as writer:
            # 保存各个工作表
            for sheet_name, data in self.results.items():
                if data:  # 只保存有数据的表
                    df = pd.DataFrame(data)
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # 创建分析汇总表
            if self.results['Summary']:
                self._create_analysis_sheet(writer)
        
        # 美化Excel格式
        self._format_excel()
        
        print(f"✓ 结果已保存到 {self.excel_path}")
        print(f"✓ 共记录 {len(self.results['Summary'])} 个实验")
    
    def _create_analysis_sheet(self, writer):
        """创建分析汇总工作表"""
        df_summary = pd.DataFrame(self.results['Summary'])
        
        # 按方法分组统计
        if not df_summary.empty:
            analysis_data = []
            
            for method in df_summary['Method'].unique():
                method_data = df_summary[df_summary['Method'] == method]
                analysis_data.append({
                    'Method': method,
                    'Num_Experiments': len(method_data),
                    'Avg_Accuracy': method_data['Accuracy'].mean(),
                    'Std_Accuracy': method_data['Accuracy'].std(),
                    'Max_Accuracy': method_data['Accuracy'].max(),
                    'Min_Accuracy': method_data['Accuracy'].min(),
                    'Avg_Training_Time': method_data['Training_Time_Sec'].mean(),
                    'Avg_Convergence_Round': method_data['Convergence_Round'].mean()
                })
            
            df_analysis = pd.DataFrame(analysis_data)
            df_analysis.to_excel(writer, sheet_name='Analysis', index=False)
    
    def _format_excel(self):
        """美化Excel格式"""
        wb = openpyxl.load_workbook(self.excel_path)
        
        # 定义样式
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 格式化每个工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # 设置表头样式
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # 设置数据样式和列宽
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 自动调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 冻结首行
            ws.freeze_panes = 'A2'
        
        wb.save(self.excel_path)
    
    def export_json(self, json_path='results/experiment_results.json'):
        """导出JSON格式的结果"""
        json_path = Path(json_path)
        json_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(self.results, f, indent=2, ensure_ascii=False)
        
        print(f"✓ JSON结果已保存到 {json_path}")


class ExperimentProgressTracker:
    """实验进度追踪器"""
    
    def __init__(self, total_experiments):
        self.total = total_experiments
        self.completed = 0
        self.start_time = datetime.now()
    
    def update(self, experiment_id):
        """更新进度"""
        self.completed += 1
        elapsed = (datetime.now() - self.start_time).total_seconds()
        avg_time = elapsed / self.completed
        remaining = (self.total - self.completed) * avg_time
        
        print(f"\n{'='*70}")
        print(f"实验进度: {self.completed}/{self.total} ({self.completed/self.total*100:.1f}%)")
        print(f"已完成: {experiment_id}")
        print(f"已用时间: {self._format_time(elapsed)}")
        print(f"预计剩余: {self._format_time(remaining)}")
        print(f"{'='*70}\n")
    
    def _format_time(self, seconds):
        """格式化时间显示"""
        hours = int(seconds // 3600)
        minutes = int((seconds % 3600) // 60)
        secs = int(seconds % 60)
        return f"{hours}h {minutes}m {secs}s"


def test_logger():
    """测试记录器"""
    logger = ExperimentLogger('results/test_results.xlsx')
    
    # 模拟记录一些结果
    for i in range(5):
        result = logger.create_result_dict(
            dataset='australian',
            partition_type='lda',
            alpha=0.1,
            num_clients=10,
            learning_rate=0.001,
            epsilon=1.0,
            method='fedavg',
            metrics={'accuracy': 0.85 + i*0.01, 'precision': 0.83, 'recall': 0.82, 'f1': 0.825, 'auc': 0.87},
            training_info={'total_rounds': 100, 'convergence_round': 85, 'training_time': 245.6, 
                          'avg_round_time': 2.45, 'final_loss': 0.32, 'gpu_used': 'Yes'},
            notes='Test experiment'
        )
        logger.log_result('A', result)
    
    logger.save_to_excel()
    print("测试完成!")


if __name__ == '__main__':
    test_logger()
